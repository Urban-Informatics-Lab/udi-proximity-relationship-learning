#Urban Data Integration(UDI) program is used to learn proximity 
#relationships between elements of different classes in a city.    

#Copyright (C) 2017-2018  Karan Gupta,Zheng Yang, Rishee Jain      

#This program is free software: you can redistribute it and/or modify 
#it under the terms of the GNU Affero General Public License as published 
#by the Free Software Foundation, either version 3 of the License, or 
#(at your option) any later version.      

#This program is distributed in the hope that it will be useful, but 
#WITHOUT ANY WARRANTY; without even the implied warranty of 
#MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU 
#Affero General Public License for more details. You should have 
#received a copy of the GNU Affero General Public License along with 
#this program.  If not, see <http://www.gnu.org/licenses/>



from __future__ import absolute_import
import sys
sys.setrecursionlimit(10000)
import os #For finding relevant files in working directory
from openpyxl import load_workbook #openpyxl is an excel file handling library in python
from sklearn.neighbors import NearestNeighbors # library to find nearest neighbors of a particular point
import math
from math import cos, sin, acos, atan2,asin
import scipy
import numpy as np
from haversine import haversine #calculating geo-distances
from vectors import Point, Vector #handling vectors
from qhull_2d import * #used as a supporting file for finding bounding box of a polygon
from min_bounding_rect import * #finding bounding box of a polygon
from supporting_functions import *#functions called throughout the code to perform repetitive operations
import pdb

w_p=[]#list to store all polygon element files
w_b=[]#list ot store all block(segments of corresponding linear element)elements files
w_l=[]#list to store all linear element files
w_pt=[]#list to sotre all point element files
l_name=[]
p_name=[]
pt_name=[]
b_name=[]

#open appropriate files
#The user must maintain the following folder structure in the current workint directory. All the files should be of .xlsx format
#~\Polygon --> store all polygon files here and each file should be named with the class of elements it is storing for example - Buildings, Ponds, Lakes
#~\Block --> store all Block files here and each file should be named with the class of corresponding linear element it is storing, for example - Roads, Pipeline
#~\Linear --> store all polygon files here and each file should be named with the class of elements it is storing for example - Buildings, Roads, Pipeline
#~\Point --> store all polygon files here and each file should be named with the class of elements it is storing for example - Trees, Light Pole, Traffic Sensor
currentdir=os.getcwd()

for file in os.listdir(currentdir+"/Polygon"):
    if file.endswith(".xlsx"):
		w_p.append(load_workbook(currentdir+"\\Polygon\\"+file))
		p_name.append(file.replace('.xlsx',''))
		
for file in os.listdir(currentdir+"/Block"):
    if file.endswith(".xlsx"):
		w_b.append(load_workbook(currentdir+"\\Block\\"+file))
		b_name.append(file.replace('.xlsx',''))
		
for file in os.listdir(currentdir+"/Point"):
    if file.endswith(".xlsx"):
		w_pt.append(load_workbook(currentdir+"\\Point\\"+file))
		pt_name.append(file.replace('.xlsx',''))
		
for file in os.listdir(currentdir+"/Linear"):
    if file.endswith(".xlsx"):
		w_l.append(load_workbook(currentdir+"\\Linear\\"+file))
		l_name.append(file.replace('.xlsx',''))

n_linear=len(w_l)
n_block=len(w_b)
n_point=len(w_pt)
n_poly=len(w_p)

#The threshold values defined by the user for the analysis. The values should be define in m unless noted otherwise
rough_filter_thresh=.002# the unit is degrees
p_p_thresh=100
p_pt_thresh=10
p_l_thresh=100
pt_pt_thresh=10
pt_l_thresh=10
l_l_thresh=50



l_vectors=[]


#Extract all linear elements and associated coordinate data into appropriate lists
for l_n in range(0,n_linear):
	sheetr=w_l[l_n].active
	r_gid=findattribute(sheetr,u'gid')
	
	road_vectors={}

	for i in range(2,sheetr.max_row + 1):# forms a dictionary to linear elements
		road_vectors[int(sheetr.cell(row=i,column=r_gid).value)]=[0,0,0,0,0,0]
		
	l_vectors.append(road_vectors)


pt_dict=[]#convert gid of a tree and associated coordinate data into appropriate lists
pt_vectors=[]
pt_coordinates=[]
pt_dist_coordinates=[]
pt_cartesian_coordinates=[]

#Extract all point elements and associated coordinate data into appropriate lists
for pt_n in range(0,n_point):
	sheett=w_pt[pt_n].active#somethign p_n###################################################
	tccount=sheett.max_column
	t_gid=findattribute(sheett,u'gid')
	
	t_lat,t_long=findlatlong(sheett)
	t_coordinates=[[1,[1,1]]]
	t_dist_coordinates=[[1,1]]
	t_cartesian_coordinates=[[1,1]]
	tree_dict={}
	tree_vectors={}
	#sheett.cell(row=1,column=tccount+1).value=u'Next_to'
	for i in range(2,sheett.max_row+1): # forms a dictionary of point elements
		
		temp2=sheett.cell(row=i,column=t_gid).value#the gid of the tree
		tree_dict[temp2]=i
		
		tree_vectors[int(sheett.cell(row=i,column=t_gid).value)]=[0,0,0,0,0,0]
		t_coordinates.append([sheett.cell(row=i,column=t_gid).value,(float(sheett.cell(row=i,column=t_lat).value),float(sheett.cell(row=i,column=t_long).value))])
		t_dist_coordinates.append([float(sheett.cell(row=i,column=t_lat).value)*float(math.pi/180),float(sheett.cell(row=i,column=t_long).value)*float(math.pi/180)])
		t_cartesian_coordinates.append(geoToCartesian(float(sheett.cell(row=i,column=t_lat).value),float(sheett.cell(row=i,column=t_long).value)))
	pt_vectors.append(tree_vectors)
	pt_coordinates.append(t_coordinates)
	pt_dist_coordinates.append(t_dist_coordinates)
	pt_cartesian_coordinates.append(t_cartesian_coordinates)
	
	pt_dict.append(tree_dict)

b_seg_dict=[]

b_seg_r_name=[]
#Extract all block elements and associated coordinate data into appropriate lists
for b_n in range(0,n_block):
	
	sheets=w_b[b_n].active#something pn####################
	s_kml_index=findattribute(sheets,u'kml')
	s_road_index=findattribute(sheets,u'Loc Street gid')
	s_gid_index=findattribute(sheets,u'gid')
	seg_dict={}
	
	seg_r_name=[]
	for i in range (2, sheets.max_row+1): # extracts the coordinates of segments of roads from the roads table
		
		temp=sheets.cell(row=i,column=s_kml_index).value
		if temp==None or temp==[]:
			continue
		start_index=temp.find(u'coordinates>')+len(u'coordinates>')
		seg_polygon=[]
		seg_polygon1=[]
		seg_c_polygon=[]
		#print temp, i, temp[start_index]
		while temp[start_index] != u'<':
			break_index=temp[start_index:].find(u",0")+start_index
			#print temp[break_index], break_index, start_index, temp[start_index:]
		
			middle_index=temp[start_index:break_index].find(u',')+start_index
			#print start_index, middle_index, break_index
			
			seg_polygon.append((float(temp[middle_index+1:break_index]),float(temp[start_index:middle_index])))
			seg_polygon1.append([float(temp[middle_index+1:break_index])*float(math.pi/180),float(temp[start_index:middle_index])*float(math.pi/180)])
			temp_lat=float(temp[middle_index+1:break_index])
			temp_long=float(temp[start_index:middle_index])
			tempx,tempy=geoToCartesian(temp_lat,temp_long)
			seg_c_polygon.append([tempx,tempy])
			start_index=break_index+2
			
		temp=sheets.cell(row=i,column=s_road_index).value.lower().strip()
		seg_r_name.append(temp)
			
		temp2=sheets.cell(row=i,column=s_gid_index).value#the gid of the segment
		try:
			temp3=seg_dict[temp]
			temp3.append([i,temp2,seg_polygon,seg_polygon1,seg_c_polygon])
			seg_dict[temp]=temp3
			
		except KeyError:
			seg_dict[temp]=[[i,temp2,seg_polygon,seg_polygon1,seg_c_polygon]]
			
	b_seg_r_name.append(list(set(seg_r_name)))
	b_seg_dict.append(seg_dict)
	
#bb_matrix=[[-1 for x in range(sheetb.max_row+1)] for y in range(sheetb.max_row+1)]
#br_matrix=[[-1 for x in range(sheetr.max_row+1)] for y in range(sheetb.max_row+1)]
#bt_matrix=[[-1 for x in range(sheett.max_row+1)] for y in range(sheetb.max_row+1)]
#tr_matrix=[[-1 for x in range(sheetr.max_row+1)] for y in range(sheett.max_row+1)]

p_roof=[]
p_roof1=[]
p_c_roof=[]
p_sides=[]

a_p_roof=[]
a_p_roof1=[]
a_p_c_roof=[]
a_p_sides=[]

p_b_c_coordinates=[]
p_b_coordinates=[]
p_b_dist_coordinates=[]
p_vectors=[]
#Extract all polygon elements and associated coordinate data into appropriate lists
for p_n in range(0,n_poly):
	sheetb=w_p[p_n].active#somethign p_n###################################################
	b_gid=findattribute(sheetb,u'gid')
	b_kml_index=findattribute(sheetb,u'KML')
	b_lat,b_long=findlatlong(sheetb)
	#b_r=findattribute(sheetb,u'street_name')
	#bccount=sheetb.max_column-3
	#sheetb.cell(row=1,column=bccount+4).value=u"Location_Road"
	
	b_coordinates=[[1,[1,1]]]
	b_dist_coordinates=[[1,1]]
	b_c_coordinates=[[1,1]]
	
	b_roof=[[[1,1],[1,1],[1,1]]]
	b_roof1=[[[1,1],[1,1],[1,1]]]
	b_cartesian_roof=[[[1,1],[1,1],[1,1]]]
	b_sides=[[1]]

	approx_b_roof=[[[1,1],[1,1],[1,1]]]
	approx_b_roof1=[[[1,1],[1,1],[1,1]]]
	approx_b_cartesian_roof=[[[1,1],[1,1],[1,1]]]
	approx_b_sides=[[1]]
	building_vectors={}
	
# next loops also finds out the sides of polygon
	for i in range(2,sheetb.max_row+1):# generates arrays of building profile and approx building profile in geo coordinates (degrees), geo coordinates (PI) and cartesian coordinates

		building_vectors[int(sheetb.cell(row=i,column=b_gid).value)]=[0,0,0,0,0,0,0,0,0]
		kml_lat=float(0)
		kml_long=float(0)
	#	print i, b_kml_index
		temp=sheetb.cell(row=i,column=b_kml_index).value
		start_index=temp.find(u'coordinates>')+len(u'coordinates>')
		polygon=[]
		polygon1=[]
		c_polygon=[]
		
		approx_polygon=[]
		approx_polygon1=[]
		approx_c_polygon=[]
		#print temp, i, temp[start_index]
		while temp[start_index] != u'<':
			break_index=temp[start_index:].find(u",0")+start_index
			#print temp[break_index], break_index, start_index, temp[start_index:]
		
			middle_index=temp[start_index:break_index].find(u',')+start_index
			#print start_index, middle_index, break_index
			
			polygon.append((float(temp[middle_index+1:break_index]),float(temp[start_index:middle_index])))
			polygon1.append([float(temp[middle_index+1:break_index])*float(math.pi/180),float(temp[start_index:middle_index])*float(math.pi/180)])
			temp_lat=float(temp[middle_index+1:break_index])*float(math.pi/180)
			temp_long=float(temp[start_index:middle_index])*float(math.pi/180)
			tempx,tempy=geoToCartesian(float(temp[middle_index+1:break_index]),float(temp[start_index:middle_index]))
			c_polygon.append([tempx,tempy])
			start_index=break_index+2
		
		#if i==4:
		#	print polygon
		#	print polygon1
		#	print c_polygon
		
		approx_polygon=polygon
		approx_polygon1=polygon1
		approx_c_polygon=c_polygon
		
		if len(c_polygon)>5:
			# Find convex hull
			xy_points=np.array(c_polygon)
			hull_points = qhull2D(xy_points)
			#print i, c_polygon
			# Reverse order of points, to match output from other qhull implementations
			hull_points = hull_points[::-1]

			#print 'Convex hull points: \n', hull_points, "\n"
			#print hull_points
			# Find minimum area bounding rectangle
			(rot_angle, area, width, height, center_point, corner_points) = minBoundingRect(hull_points)
			approx_c_polygon=corner_points.tolist()
			#print i, c_polygon
			temp_min_polygon=[]
			temp_min_polygon1=[]
			for p in range(0,len(corner_points)):
				temp_min_polygon.append(cartesianToGeo(c_polygon[p][0],c_polygon[p][1]))
				temp_xx,temp_yy=cartesianToGeo(c_polygon[p][0],c_polygon[p][1])
				temp_min_polygon1.append([temp_xx*float(math.pi/180),temp_yy*float(math.pi/180)])
			approx_polygon=temp_min_polygon
		
			#print polygon
			approx_polygon1=temp_min_polygon1
			#print polygon1
			#time.sleep(5)
		del polygon[-1]
		del polygon1[-1]
		del c_polygon[-1]
		
		#del approx_polygon[-1]
		#del approx_polygon1[-1]
		#del approx_c_polygon[-1]
		#if i==2:
		#	print polygon
		#	print approx_polygon
		#	print approx_polygon1
		#	print approx_c_polygon
		
		b_roof.append(polygon)
		b_roof1.append(polygon1)
		b_cartesian_roof.append(c_polygon)
		b_sides.append(findSides(c_polygon))
		
		approx_b_roof.append(approx_polygon)
		approx_b_roof1.append(approx_polygon1)
		approx_b_cartesian_roof.append(approx_c_polygon)
		approx_b_sides.append(findSides(approx_c_polygon))
		
		
		kml_long=sum(list(zip(*polygon)[1]))/len(polygon)
		kml_lat=sum(list(zip(*polygon)[0]))/len(polygon)
		
		b_coordinates.append([sheetb.cell(row=i,column=b_gid).value,(float(kml_lat),float(kml_long))])
		b_dist_coordinates.append([float(kml_lat)*float(math.pi/180),float(kml_long)*float(math.pi/180)])
		
		temp_xx,temp_yy=geoToCartesian(float(kml_lat),float(kml_long))
		b_c_coordinates.append([temp_xx,temp_yy])
		
		#temp1=sheetb.cell(row=i,column=findattribute(sheetb,u'street_name')).value.lower().strip()
		
		#p1=temp1.find(u' ')
		#p2=temp1[p1+1:].find(' ')
		#if p2==-1:
		#	temp1=temp1[:p1+2]
		#else:
		#	temp1=temp1[:p1+p2+3]
		#temp1=temp1[:p1+2]
		
		##try:
		##	sheetb.cell(row=i,column=bccount+4).value=u'Road'+str(road_dict[temp1])
		##except KeyError:
		##	print u'road:', temp1, u' not found in building ', i

	p_roof.append(b_roof)
	p_roof1.append(b_roof1)
	p_c_roof.append(b_cartesian_roof)
	p_sides.append(b_sides)

	a_p_roof.append(approx_b_roof)
	a_p_roof1.append(approx_b_roof1)
	a_p_c_roof.append(approx_b_cartesian_roof)
	a_p_sides.append(approx_b_sides)
	
	p_b_coordinates.append(b_coordinates)
	p_b_dist_coordinates.append(b_dist_coordinates)
	p_b_c_coordinates.append(b_c_coordinates)
	p_vectors.append(building_vectors)

print "done buildings"
print "done till before buildings"


p_final_building = {}

p_final_building_roads={}
p_final_intersection= {}
#Learning relations with polygon elements
for p_n in range(0,n_poly):

	sheetb=w_p[p_n].active#########
	b_roof=p_roof[p_n]
	b_roof1=p_roof1[p_n]
	b_carteisan_roof=p_c_roof[p_n]
	
	b_sides=p_sides[p_n]

	approx_b_roof=a_p_roof[p_n]
	approx_b_roof1=a_p_roof1[p_n]
	approx_b_cartesian_roof=a_p_c_roof[p_n]
	approx_b_sides=a_p_sides[p_n]
	
	b_coordinates=p_b_coordinates[p_n]
	b_dist_coordinates=p_b_dist_coordinates[p_n]
	b_c_coordinates=p_b_c_coordinates[p_n]
	building_vectors=p_vectors[p_n]
	final_building = {}
	
	final_building_roads={}
	final_intersection= {}
	for i in range(2,sheetb.max_row+1):
	#try:print
	
		i_valid=[None,50]
		
		print i
		#pdb.set_trace()
		temp_roof=b_roof[i-1]
		approx_temp_roof=approx_b_roof[i-1]
		
		lat_list=list(zip(*temp_roof)[0])
		long_list=list(zip(*temp_roof)[1])
		
		max_lat_allowed=max(lat_list)+float(rough_filter_thresh)
		min_lat_allowed=min(lat_list)-float(rough_filter_thresh)
		max_long_allowed=max(long_list)+float(rough_filter_thresh)
		min_long_allowed=min(long_list)-float(rough_filter_thresh)
		
			
		#print b_roof1[i-1]
		temp_roof1=b_roof1[i-1]
		temp_roof1 = np.array(temp_roof1)
		c_temp_roof1=b_cartesian_roof[i-1]
		approx_temp_roof1=approx_b_roof1[i-1]
		approx_temp_roof1 = np.array(approx_temp_roof1)
		approx_c_temp_roof1=approx_b_cartesian_roof[i-1]
		
		distances=[]
		indices=[]
				
		################### relation between polygon(building) and linear (road) elements
		for b_n in range(0,n_block):
			
			building_roads=b_seg_r_name[b_n]
			seg_dict=b_seg_dict[b_n]
			#road_dict=l_dict[b_n]
			road_vectors=l_vectors[b_n]
			sheets=w_b[b_n].active##
			s_gid_index=findattribute(sheets,u'gid')
		
			
				
			for l in range(0,len(building_roads)):############################################################################################
	#if i==4:
			#	print "Length building roads= ",len(building_roads),l, building_roads
				jk=0
				road_segments=[]
				try:
				
					road_segments=seg_dict[building_roads[l]] #[[i,temp2,seg_polygon,seg_polygon1,seg_c_polygon]]
				except KeyError:
					print "Segment of ", building_roads[l], " not avalable for buildings"
					
					continue
				#	print "Length building roads= ",len(building_roads),l, building_roads
				 #[[i,temp2,seg_polygon,seg_polygon1,seg_c_polygon]]
				
				seg_side=[]
				ssi=[]
				a_index=[]
				af_index=[]
				ab_index=[]
				
				approx_a_index=[]
				approx_af_index=[]
				approx_ab_index=[]
				
				min_distance_segment=p_l_thresh
				s_segment=[]
		
				for n in range(0,len(road_segments)):
					
					segment=road_segments[n]
			
					for si in range(0,len(segment[4])-1):
						
						if segment[2][si][0] <= max_lat_allowed:
				
							if segment[2][si][0]>= min_lat_allowed:
								if segment[2][si][1]<= max_long_allowed:
									if segment[2][si][1]>= min_long_allowed:
										
										
										if isInternal(b_c_coordinates[i-1],segment[4][si],segment[4][si+1]):
											
											t_seg_side=defineSideOfSegment(b_c_coordinates[i-1],segment[4][si],segment[4][si+1])
											
											#perDistancePointLine_2(b_c_coordinates[i-1],segment[4][si],segment[4][si+1])
											
											t_a_index=findNearestRoofPoint(c_temp_roof1,segment[4][si],segment[4][si+1])
											
											t_af_index,t_ab_index=findAdjacentIndex(c_temp_roof1, t_a_index)
											
											s_point1,s_point2=distanceBetweenSegmentandBuilding(c_temp_roof1[t_ab_index],c_temp_roof1[t_a_index],c_temp_roof1[t_af_index],segment[4][si],segment[4][si+1])
											
											approx_t_seg_side=defineSideOfSegment(b_c_coordinates[i-1],segment[4][si],segment[4][si+1])
											
											#perDistancePointLine_2(b_c_coordinates[i-1],segment[4][si],segment[4][si+1])
											
											approx_t_a_index=findNearestRoofPoint(approx_c_temp_roof1,segment[4][si],segment[4][si+1])
											
											approx_t_af_index,approx_t_ab_index=findAdjacentIndex(approx_c_temp_roof1, approx_t_a_index)
											
											approx_s_point1,approx_s_point2=distanceBetweenSegmentandBuilding(approx_c_temp_roof1[approx_t_ab_index],approx_c_temp_roof1[approx_t_a_index],approx_c_temp_roof1[approx_t_af_index],segment[4][si],segment[4][si+1])
											
											
											s_c_point1=cartesianToGeo(s_point1.x,s_point1.y)
											s_c_point2=cartesianToGeo(s_point2.x,s_point2.y)
											temp_min_distance_segment=haversine(s_c_point1,s_c_point2)*1000
											
											
											#print temp_min_distance_segment
											
											if min_distance_segment>temp_min_distance_segment:
												min_distance_segment=temp_min_distance_segment
												a_index=t_a_index
												ab_index=t_ab_index
												af_index=t_af_index
												approx_a_index=approx_t_a_index
												approx_ab_index=approx_t_ab_index
												approx_af_index=approx_t_af_index
												seg_side=t_seg_side
												ssi=si
												s_segment=segment
												#pdb.set_trace()
												#print "buidling ", i, " road ", n, si
												#print building_roads, min_distance_segment
				flags1=0
				flagsb=0
				
				O1=[]
				O2=[]
				goto=0
				try:
					#pdb.set_trace()
					s_segment[4][ssi]
					s_segment[4][ssi+1]
					approx_c_temp_roof1[approx_a_index]
					approx_c_temp_roof1[approx_ab_index]
					O1=findOverlap([s_segment[4][ssi],s_segment[4][ssi+1]],approx_c_temp_roof1[approx_ab_index],approx_c_temp_roof1[approx_a_index])
					O2=findOverlap([s_segment[4][ssi],s_segment[4][ssi+1]],approx_c_temp_roof1[approx_a_index],approx_c_temp_roof1[approx_af_index])
				except ZeroDivisionError :
					print "Building ", i ,"or  road ", s_segment[0], s_segment[1] , "is bad"
					goto=2
				except IndexError :
					goto=2
					
				s_building_side=[]
				s_building_side1=[]
				#print "inside building roads", O1, O2
				if goto==0:
					if O1==0 and O2==0:
						flags1=1
					else:
						
						if O1>O2:
							s_building_side=extractSide(approx_b_sides[i-1],approx_ab_index,approx_a_index)
						elif O1<O2:
							s_building_side=extractSide(approx_b_sides[i-1],approx_a_index,approx_af_index)
						else: 
							s_building_side=extractSide(approx_b_sides[i-1],approx_ab_index,approx_a_index)
							s_building_side1=extractSide(approx_b_sides[i-1],approx_a_index,approx_af_index)
							flagsb=1
						
						#if i==4:
						#	print s_building_side,s_building_side1, i, l
									

					if flags1==0:
					
						temp_vector=building_vectors[int(b_coordinates[i-1][0])]
						
						rand_1=temp_vector[0]
						rand_2=temp_vector[1]
						rand_3=temp_vector[2]
						rand_7=temp_vector[6]
						rand_8=temp_vector[7]
						rand_9=temp_vector[8]
						n_pre=temp_vector[3]
						ave_pre=temp_vector[4]
						std_pre=temp_vector[5]
						n_new=n_pre+1
						ave_new=float((ave_pre*n_pre+min_distance_segment)/float(n_pre+1))
						sumx=ave_pre*n_pre
						sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
						std_new=(float((float(n_new)*(min_distance_segment**2+sumx2)-(min_distance_segment+sumx)**2)/float((n_new)**2)))**.5
						building_vectors[int(b_coordinates[i-1][0])]=[rand_1,rand_2,rand_3,n_new,ave_new,std_new,rand_7,rand_8,rand_9]
						
						temp_vector=road_vectors[int(building_roads[l])]
						
						rand_4=temp_vector[3]
						rand_5=temp_vector[4]
						rand_6=temp_vector[5]
					
						n_pre=temp_vector[0]
						ave_pre=temp_vector[1]
						std_pre=temp_vector[2]
						n_new=n_pre+1
						
						ave_new=float((ave_pre*n_pre+min_distance_segment)/float(n_pre+1))
						sumx=ave_pre*n_pre
						sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
						std_new=(float((float(n_new)*(min_distance_segment**2+sumx2)-(min_distance_segment+sumx)**2)/float((n_new)**2)))**.5
							
						road_vectors[int(building_roads[l])]=[n_new,ave_new,std_new,rand_4,rand_5,rand_6]
							
						try:
							final_building_roads=p_final_building_roads[p_n]
							temp_final_building=final_building_roads[i]
							#if i==4: 
					#			print "try", final_building_roads[i]
							
							if flagsb==1:
								temp_final_building.append([b_name[b_n]+str('_')+str(building_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Side_"+str(s_building_side),"Seg_Side_"+str(seg_side)])
								temp_final_building.append([b_name[b_n]+str('_')+str(building_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Side_"+str(s_building_side1),"Seg_Side_"+str(seg_side)])
								#br_matrix[i][road_dict_2[road_dict[building_roads[l]]]]=min_distance_segment
								
							else: 
								temp_final_building.append([b_name[b_n]+str('_')+str(building_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Side_"+str(s_building_side),"Seg_Side_"+str(seg_side)])
								#br_matrix[i][road_dict_2[road_dict[building_roads[l]]]]=min_distance_segment
							final_building_roads[i]=temp_final_building
							p_final_building_roads[p_n]=final_building_roads
							
						except KeyError:
							temp_final_building=[]
							if flagsb==1:
								
								temp_final_building.append([b_name[b_n]+str('_')+str(building_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Side_"+str(s_building_side),"Seg_Side_"+str(seg_side)])
								temp_final_building.append([b_name[b_n]+str('_')+str(building_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Side_"+str(s_building_side1),"Seg_Side_"+str(seg_side)])
								#br_matrix[i][road_dict_2[road_dict[building_roads[l]]]]=min_distance_segment
							else:
								#pdb.set_trace()
								temp_final_building.append([b_name[b_n]+str('_')+str(building_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Side_"+str(s_building_side),"Seg_Side_"+str(seg_side)])
								#br_matrix[i][road_dict_2[road_dict[building_roads[l]]]]=min_distance_segment
							final_building_roads[i]=temp_final_building
							p_final_building_roads[p_n]=final_building_roads
							
		################### relation between polygon(building) and polygon (building) elements
		for p_n_j in range(p_n,n_poly):
			sheetbj=w_p[p_n_j].active########
			b_roof_j=p_roof[p_n_j]
			b_roof1_j=p_roof1[p_n_j]
			b_cartesian_roof_j=p_c_roof[p_n_j]
			b_sides_j=p_sides[p_n_j]

			approx_b_roof_j=a_p_roof[p_n_j]
			approx_b_roof1_j=a_p_roof1[p_n_j]
			approx_b_cartesian_roof_j=a_p_c_roof[p_n_j]
			approx_b_sides_j=a_p_sides[p_n_j]
			
			b_coordinates_j=p_b_coordinates[p_n_j]
			b_dist_coordinates_j=p_b_dist_coordinates[p_n_j]
			b_c_coordinates_j=p_b_c_coordinates[p_n_j]
			building_vectors_j=p_vectors[p_n_j]
			b_kml_index_j=findattribute(sheetbj,u'KML')
			p_s_index=2
			if p_n==p_n_j:
				p_s_index=i+1
			final_building_j={}
			if p_s_index>sheetbj.max_row:
				continue
			for j in range(p_s_index,sheetbj.max_row+1):
				l_lat=list(zip(*b_roof_j[j-1])[0])
				l_long=list(zip(*b_roof_j[j-1])[1])
				if min(l_lat) <= max_lat_allowed:
					if max(l_lat)>= min_lat_allowed:
						if min(l_long)<= max_long_allowed:
							if max(l_long)>= min_long_allowed:
								temp=sheetbj.cell(row=j,column=b_kml_index_j).value
								#start_index=temp.find(coordinates)+len(coordinates)
								
							
								Y = np.array(list(temp_roof1))
								nbrs = NearestNeighbors(n_neighbors=1, algorithm='ball_tree',metric=u'haversine').fit(b_roof1_j[j-1])
								distances, indices=nbrs.kneighbors(Y)
								distances=distances.tolist()
								
								approx_Y = np.array(list(approx_temp_roof1))
								approx_nbrs = NearestNeighbors(n_neighbors=1, algorithm='ball_tree',metric=u'haversine').fit(approx_b_roof1_j[j-1])
								approx_distances, approx_indices=approx_nbrs.kneighbors(approx_Y)
								approx_distances=approx_distances.tolist()
								
								dists=[]
								for m in range(0,len(distances)):
									dists.append(distances[m][0]*1000*6371)
								distances=dists
								
								if min(distances)<p_p_thresh :
									
									b_temp_kk=list(b_cartesian_roof_j[j-1])
									a_index=distances.index(min(distances))
									b_index=indices[a_index][0]
									af_index,ab_index=findAdjacentIndex(c_temp_roof1, a_index)
									bf_index,bb_index=findAdjacentIndex(b_temp_kk,b_index)
									
									approx_b_temp_kk=approx_b_cartesian_roof_j[j-1]
									approx_a_index=approx_distances.index(min(approx_distances))
									approx_b_index=approx_indices[approx_a_index][0]
									#print approx_b_index
									approx_af_index,approx_ab_index=findAdjacentIndex(approx_c_temp_roof1, approx_a_index)
									approx_bf_index,approx_bb_index=findAdjacentIndex(approx_b_temp_kk,approx_b_index)
									
									min_distance=[]
									try:
									
										min_distance=findDistanceBetweenSegments(c_temp_roof1[af_index],c_temp_roof1[a_index],c_temp_roof1[ab_index],b_temp_kk[bf_index],b_temp_kk[b_index],b_temp_kk[bf_index])
									except Exception as e:
										print e
										#pdb.set_trace()
															
									flaga=0
									flagb=0
									a_building_side=[]
									a_building_side1=[]
									
									b_building_side=[]
									b_building_side1=[]
									flag1=0
									flag2=0
								
							
									
									#print approx_bb_index, j, approx_bb_index, "DD", approx_b_temp_kk, approx_b_index
									O1=[]
									O2=[]
									try:
									
											
										O1=findOverlap(approx_c_temp_roof1,approx_b_temp_kk[approx_bb_index],approx_b_temp_kk[approx_b_index])
										O2=findOverlap(approx_c_temp_roof1,approx_b_temp_kk[approx_b_index],approx_b_temp_kk[approx_bf_index])
									except ZeroDivisionError:
										print "Building ", i ,"or ", j , "is bad"
										continue
										
									#print j, "j", approx_b_temp_kk, approx_bb_index
									if O1==0 and O2==0:
										flag1=1
									else:
										#if i==2:
										#	print "O1,O2",O1,O2
										if O1>O2:
											b_building_side=extractSide(approx_b_sides_j[j-1],approx_bb_index,approx_b_index)
										elif O2>O1:
											b_building_side=extractSide(approx_b_sides_j[j-1],approx_b_index,approx_bf_index)
										else: 
											b_building_side=extractSide(approx_b_sides_j[j-1],approx_bb_index,approx_b_index)
											b_building_side1=extractSide(approx_b_sides_j[j-1],approx_b_index,approx_bf_index)
											flagb=1
											
									try:
										O3=findOverlap(approx_b_temp_kk,approx_c_temp_roof1[approx_ab_index],approx_c_temp_roof1[approx_a_index])
										O4=findOverlap(approx_b_temp_kk,approx_c_temp_roof1[approx_a_index],approx_c_temp_roof1[approx_af_index])
									except ZeroDivisionError:
										print "Building ", i ,"or ", j , "is bad"
										continue
									if O3==0 and O4==0:
										flag2=1
									else:
										#if i==2:
										#	print "O3,O4",O3,O4
										if O3>O4:
											a_building_side=extractSide(approx_b_sides[i-1],approx_ab_index,approx_a_index)
										elif O3<O4:
											a_building_side=extractSide(approx_b_sides[i-1],approx_a_index,approx_af_index)
										else: 
											a_building_side=extractSide(approx_b_sides[i-1],approx_ab_index,approx_a_index)
											a_building_side1=extractSide(approx_b_sides[i-1],approx_a_index,approx_af_index)
											flaga=1
																	
												
									if flag2==0:
										temp_vector=building_vectors[int(b_coordinates[i-1][0])]
										rand_4=temp_vector[3]
										rand_5=temp_vector[4]
										rand_6=temp_vector[5]
										rand_7=temp_vector[6]
										rand_8=temp_vector[7]
										rand_9=temp_vector[8]
										n_pre=temp_vector[0]
										ave_pre=temp_vector[1]
										std_pre=temp_vector[2]
										n_new=n_pre+1
										ave_new=float((ave_pre*n_pre+min_distance)/float(n_pre+1))
										sumx=ave_pre*n_pre
										sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
										std_new=(float((float(n_new)*(min_distance**2+sumx2)-(min_distance+sumx)**2)/float((n_new)**2)))**.5
										building_vectors[int(b_coordinates[i-1][0])]=[n_new,ave_new,std_new,rand_4,rand_5,rand_6,rand_7,rand_8,rand_9]
										
																				
										try:
											final_building=p_final_building[p_n]
											temp_final_building=final_building[i]
											if flaga==1:
												temp_final_building.append([p_name[p_n_j],str(b_coordinates_j[j-1][0]),min_distance,"Side_"+str(a_building_side)])
												temp_final_building.append([p_name[p_n_j],str(b_coordinates_j[j-1][0]),min_distance,"Side_"+str(a_building_side1)])
												#bb_matrix[i][j]=min_distance
										
											else: 
												temp_final_building.append([p_name[p_n_j],str(b_coordinates_j[j-1][0]),min_distance,"Side_"+str(a_building_side)])
												#bb_matrix[i][j]=min_distance
										
												
											final_building[i]=temp_final_building
											p_final_building[p_n]=final_building
										except KeyError:
											temp_final_building=[]
											if flaga==1:
												temp_final_building.append([p_name[p_n_j],str(b_coordinates_j[j-1][0]),min_distance,"Side_"+str(a_building_side)])
												temp_final_building.append([p_name[p_n_j],str(b_coordinates_j[j-1][0]),min_distance,"Side_"+str(a_building_side1)])
												#bb_matrix[i][j]=min_distance
										
											else:
												temp_final_building.append([p_name[p_n_j],str(b_coordinates_j[j-1][0]),min_distance,"Side_"+str(a_building_side)])
												#bb_matrix[i][j]=min_distance
										
											final_building[i]=temp_final_building
											p_final_building[p_n]=final_building
									
										
									if flag1==0:
									
										temp_vector=building_vectors[int(b_coordinates_j[j-1][0])]
										rand_4=temp_vector[3]
										rand_5=temp_vector[4]
										rand_6=temp_vector[5]
										rand_7=temp_vector[6]
										rand_8=temp_vector[7]
										rand_9=temp_vector[8]
										n_pre=temp_vector[0]
										ave_pre=temp_vector[1]
										std_pre=temp_vector[2]
										n_new=n_pre+1
										ave_new=float((ave_pre*n_pre+min_distance)/float(n_pre+1))
										sumx=ave_pre*n_pre
										sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
										std_new=(float((float(n_new)*(min_distance**2+sumx2)-(min_distance+sumx)**2)/float((n_new)**2)))**.5
										building_vectors[int(b_coordinates_j[j-1][0])]=[n_new,ave_new,std_new,rand_4,rand_5,rand_6,rand_7,rand_8,rand_9]
										try:
											final_building_j=p_final_building[p_n_j]
											temp_final_building=final_building_j[j]
											if flagb==1:
												temp_final_building.append([p_name[p_n],str(b_coordinates[i-1][0]),min_distance,"Side_"+str(b_building_side)])
												temp_final_building.append([p_name[p_n],str(b_coordinates[i-1][0]),min_distance,"Side_"+str(b_building_side1)])
												#bb_matrix[j][i]=min_distance
										#		if i==2:
										#			print "added two-in j"
											
											else:
												temp_final_building.append([p_name[p_n],str(b_coordinates[i-1][0]),min_distance,"Side_"+str(b_building_side)])
												#bb_matrix[j][i]=min_distance
										#		if i==2:
										#			print "added two-in j"
											final_building_j[j]=temp_final_building
											p_final_building[p_n_j]=final_building_j
										except KeyError:
											temp_final_building=[]
											if flagb==1:
												temp_final_building.append([p_name[p_n],str(b_coordinates[i-1][0]),min_distance,"Side_"+str(b_building_side)])
												
												temp_final_building.append([p_name[p_n],str(b_coordinates[i-1][0]),min_distance,"Side_"+str(b_building_side1)])
												#bb_matrix[j][i]=min_distance
												#bb_matrix
										#		if i==2:
										#			print "added two-except in j"
											else:
												temp_final_building.append([p_name[p_n],str(b_coordinates[i-1][0]),min_distance,"Side_"+str(b_building_side)])
												#bb_matrix[j][i]=min_distance
										#		if i==2:	
										#			print "added one-except in j"
											final_building_j[j]=temp_final_building
											p_final_building[p_n_j]=final_building
										#pdb.set_trace()
										#if i==2:
											#	print final_building[j],"i=",i,"j=",j
										
		################### relation between polygon(building) and point (tree) elements
		for pt_n in range(0,n_point):
			sheett=w_pt[pt_n].active####
			tree_vectors=pt_vectors[pt_n]
			t_coordinates=pt_coordinates[pt_n]
			t_dist_coordinates=pt_dist_coordinates[pt_n]
			t_cartesian_coordinates=pt_cartesian_coordinates[pt_n]
			
			pt_dict.append(tree_dict)
			potential_tree=[]
			for j in range(2,sheett.max_row+1):
				if t_coordinates[j-1][1][0] <= max_lat_allowed:
				
					if t_coordinates[j-1][1][0]>= min_lat_allowed:
						if t_coordinates[j-1][1][1]<= max_long_allowed:
							if t_coordinates[j-1][1][1]>= min_long_allowed:
								#print u'here', j
								potential_tree.append([j,t_dist_coordinates[j-1]])
			if potential_tree !=[]:
				X = np.array(list(zip(*potential_tree)[1]))
				nbrs = NearestNeighbors(n_neighbors=1, algorithm='ball_tree',metric=u'haversine').fit(b_roof1[i-1])
				#result=nbrs.radius_neighbors(X,.010/6371.0,return_distance=True)
				
				
				#distances=np.asarray(result[0])
				#indices=np.asarray(result[1])#indexes of intersections near to a building
				
				distances, indices=nbrs.kneighbors(X)
				distances=distances*1000*6371
				
				approx_nbrs = NearestNeighbors(n_neighbors=1, algorithm='ball_tree',metric=u'haversine').fit(approx_b_roof1[i-1])
				#result=nbrs.radius_neighbors(X,.010/6371.0,return_distance=True)
				
				
				#distances=np.asarray(result[0])
				#indices=np.asarray(result[1])#indexes of intersections near to a building
				
				approx_distances, approx_indices=approx_nbrs.kneighbors(X)
			
				approx_distances=approx_distances*1000*6371
				
				#indices=np.array(indices.tolist())
				temp_c_roof=b_cartesian_roof[i-1]
				approx_temp_c_roof=approx_b_cartesian_roof[i-1]
			
				#distances=(np.array(distances.tolist()))*1000*6371
				for jt in range(0,len(indices)):
					if distances[jt][0]<p_pt_thresh*10:
						
						tree_x=t_cartesian_coordinates[potential_tree[jt][0]-1][0]
						tree_y=t_cartesian_coordinates[potential_tree[jt][0]-1][1]
						
						a_index=temp_c_roof.index(temp_c_roof[indices[jt][0]])
						af_index,ab_index=findAdjacentIndex(temp_c_roof, a_index)
						temp_result=[]
						V_line_segment=Vector.from_points(Point(temp_c_roof[ab_index][0],temp_c_roof[ab_index][1],0),Point(temp_c_roof[a_index][0],temp_c_roof[a_index][1],0))
						temp_result.append(distancePointSegment(Point(tree_x,tree_y,0),V_line_segment,Point(temp_c_roof[ab_index][0],temp_c_roof[ab_index][1],0),Point(temp_c_roof[a_index][0],temp_c_roof[a_index][1],0)))
						
						V_line_segment=Vector.from_points(Point(temp_c_roof[a_index][0],temp_c_roof[a_index][1],0),Point(temp_c_roof[af_index][0],temp_c_roof[af_index][1],0))
						temp_result.append(distancePointSegment(Point(tree_x,tree_y,0),V_line_segment,Point(temp_c_roof[a_index][0],temp_c_roof[a_index][1],0),Point(temp_c_roof[af_index][0],temp_c_roof[af_index][1],0)))
						
						approx_a_index=approx_temp_c_roof.index(approx_temp_c_roof[approx_indices[jt][0]])
						approx_af_index,approx_ab_index=findAdjacentIndex(approx_temp_c_roof, approx_a_index)
						approx_temp_result=[]
						approx_V_line_segment=Vector.from_points(Point(approx_temp_c_roof[approx_ab_index][0],approx_temp_c_roof[approx_ab_index][1],0),Point(approx_temp_c_roof[approx_a_index][0],approx_temp_c_roof[approx_a_index][1],0))
						approx_temp_result.append(distancePointSegment(Point(tree_x,tree_y,0),approx_V_line_segment,Point(approx_temp_c_roof[approx_ab_index][0],approx_temp_c_roof[approx_ab_index][1],0),Point(approx_temp_c_roof[approx_a_index][0],approx_temp_c_roof[approx_a_index][1],0)))
						
						approx_V_line_segment=Vector.from_points(Point(approx_temp_c_roof[approx_a_index][0],approx_temp_c_roof[approx_a_index][1],0),Point(approx_temp_c_roof[approx_af_index][0],approx_temp_c_roof[approx_af_index][1],0))
						approx_temp_result.append(distancePointSegment(Point(tree_x,tree_y,0),approx_V_line_segment,Point(approx_temp_c_roof[approx_a_index][0],approx_temp_c_roof[approx_a_index][1],0),Point(approx_temp_c_roof[approx_af_index][0],approx_temp_c_roof[approx_af_index][1],0)))
						
						approx_min_tree_distance_1=haversine(cartesianToGeo(tree_x,tree_y),cartesianToGeo(approx_temp_result[1][1][0],approx_temp_result[1][1][1]))*1000
						approx_min_tree_distance_2=haversine(cartesianToGeo(tree_x,tree_y),cartesianToGeo(approx_temp_result[0][1][0],approx_temp_result[0][1][1]))*1000
						
						#for k in range(0, len(indices[jt])):
						a_building_side=[]
						min_tree_distance=[]
						local_flag=0
						if temp_result[0][0]>temp_result[1][0]:
							min_tree_distance=haversine(cartesianToGeo(tree_x,tree_y),cartesianToGeo(temp_result[1][1][0],temp_result[1][1][1]))*1000
							
							if min_tree_distance<p_pt_thresh:
								if approx_min_tree_distance_1>approx_min_tree_distance_2:
								
									a_building_side=extractSide(approx_b_sides[i-1],approx_a_index,approx_af_index)
									local_flag=1
								else:
									a_building_side=extractSide(approx_b_sides[i-1],approx_ab_index,approx_a_index)
									local_flag=1
							
						else:
							min_tree_distance=haversine(cartesianToGeo(tree_x,tree_y),cartesianToGeo(temp_result[0][1][0],temp_result[0][1][1]))*1000
							if min_tree_distance<10:
								if approx_min_tree_distance_1>approx_min_tree_distance_2:
								
									a_building_side=extractSide(approx_b_sides[i-1],approx_a_index,approx_af_index)
									local_flag=1
								else:
									a_building_side=extractSide(approx_b_sides[i-1],approx_ab_index,approx_a_index)
									local_flag=1
							
							
						if local_flag==1:
						
							temp_vector=building_vectors[int(b_coordinates[i-1][0])]
							rand_1=temp_vector[0]
							rand_2=temp_vector[1]
							rand_3=temp_vector[2]
							rand_4=temp_vector[3]
							rand_5=temp_vector[4]
							rand_6=temp_vector[5]
							n_pre=temp_vector[6]
							ave_pre=temp_vector[7]
							std_pre=temp_vector[8]
							n_new=n_pre+1
							ave_new=float((ave_pre*n_pre+min_tree_distance)/float(n_pre+1))
							sumx=ave_pre*n_pre
							sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
							std_new=(float((float(n_new)*(min_tree_distance**2+sumx2)-(min_tree_distance+sumx)**2)/float((n_new)**2)))**.5
							building_vectors[int(b_coordinates[i-1][0])]=[rand_1,rand_2,rand_3,rand_4,rand_5,rand_6,n_new,ave_new,std_new]
							
							temp_vector=tree_vectors[int(t_coordinates[potential_tree[jt][0]-1][0])]
						
							rand_4=temp_vector[3]
							rand_5=temp_vector[4]
							rand_6=temp_vector[5]
							n_pre=temp_vector[0]
							ave_pre=temp_vector[1]
							std_pre=temp_vector[2]
							n_new=n_pre+1
							ave_new=float((ave_pre*n_pre+min_tree_distance)/float(n_pre+1))
							sumx=ave_pre*n_pre
							sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
							std_new=(float((float(n_new)*(min_tree_distance**2+sumx2)-(min_tree_distance+sumx)**2)/float((n_new)**2)))**.5
							tree_vectors[int(t_coordinates[potential_tree[jt][0]-1][0])]=[n_new,ave_new,std_new,rand_4,rand_5,rand_6]
							
							try:
								final_building=p_final_building[p_n]
								temp_final_building=final_building[i]
								temp_final_building.append([pt_name[pt_n],str(t_coordinates[potential_tree[jt][0]-1][0]),min_tree_distance,"Side_"+str(a_building_side)])
								final_building[i]=temp_final_building
								#bt_matrix[i][tree_dict[t_coordinates[potential_tree[jt][0]-1][0]]]=min_tree_distance
								p_final_building[p_n]=final_building
							except KeyError:
								temp_final_building=[]
								temp_final_building.append([pt_name[pt_n],str(t_coordinates[potential_tree[jt][0]-1][0]),min_tree_distance,"Side_"+str(a_building_side)])
								#bt_matrix[i][tree_dict[t_coordinates[potential_tree[jt][0]-1][0]]]=min_tree_distance
								final_building[i]=temp_final_building
								p_final_building[p_n]=final_building
			#j###################################################################################################################		
		
		#print u'Done Buildings ',int(float(i)/float(sheetb.max_row+1)*100),u'%'
	#except Exception as e:
	#	print e
	#	pdb.pm()
	
	

print "completed buildings"

#Storing learned relations of polygon in workbook objects. The files are stored in working directory
for p_n in range(0,n_poly):
	sheetb=w_p[p_n].active######
	final_building=p_final_building[p_n]
	bccount=sheetb.max_column
	for i in range(2,sheetb.max_row+1):
		
		try:
			temp=final_building[i]
			#print temp
			#temp=np.array(temp)
			#temp=temp[temp[:,2].argsort()]
			for pp in range (0,len(temp)):#([u'tree',u'Tree'+str(t_coordinates[potential_tree[j][0]-1][0]),min_tree_distance,"Side_"+str(a_building_side)])
				sheetb.cell(row=1,column=bccount+4*(pp+1)-3).value=u'Object_Type'
				sheetb.cell(row=1,column=bccount+4*(pp+1)-2).value=u'Object_Name'
				sheetb.cell(row=1,column=bccount+4*(pp+1)-1).value=u'Object_Distance'	
				sheetb.cell(row=1,column=bccount+4*(pp+1)).value=u'Parent_Side'
				sheetb.cell(row=i,column=bccount+4*(pp+1)-3).value=temp[pp][0]
				sheetb.cell(row=i,column=bccount+4*(pp+1)-2).value=temp[pp][1]
				sheetb.cell(row=i,column=bccount+4*(pp+1)-1).value=temp[pp][2]
				sheetb.cell(row=i,column=bccount+4*(pp+1)).value=temp[pp][3]
		except KeyError:
			waste=1


for p_n in range(0,n_poly):
	sheetb=w_p[p_n].active######
	final_building_roads=p_final_building_roads[p_n]
	bccount=sheetb.max_column

	for i in range(2,sheetb.max_row+1):
		try:
			temp=final_building_roads[i]
			#print temp
			#temp=np.array(temp)
			#temp=temp[temp[:,2].argsort()]
			for pp in range (0,len(temp)):#([u'tree',u'Tree'+str(t_coordinates[potential_tree[j][0]-1][0]),min_tree_distance,"Side_"+str(a_building_side)])
				sheetb.cell(row=1,column=bccount+5*(pp+1)-4).value=u'Object_Type'
				sheetb.cell(row=1,column=bccount+5*(pp+1)-3).value=u'Object_Name'
				sheetb.cell(row=1,column=bccount+5*(pp+1)-2).value=u'Object_Distance'	
				sheetb.cell(row=1,column=bccount+5*(pp+1)-1).value=u'Parent_Side'
				sheetb.cell(row=1,column=bccount+5*(pp+1)).value=u'Seg_Side'
				sheetb.cell(row=i,column=bccount+5*(pp+1)-4).value=temp[pp][0]
				sheetb.cell(row=i,column=bccount+5*(pp+1)-3).value=temp[pp][1]
				sheetb.cell(row=i,column=bccount+5*(pp+1)-2).value=temp[pp][2]
				sheetb.cell(row=i,column=bccount+5*(pp+1)-1).value=temp[pp][3]
				sheetb.cell(row=i,column=bccount+5*(pp+1)).value=temp[pp][4]
		except KeyError:
			waste=1

#code for storing workbook objects in disk. Each file is named with the corresponding element class (building, pond, lake) follwed by "-result"
while True:
	try:
		for p_n in range(0,n_poly):
			w_p[p_n].save(p_name[p_n]+'-result.xlsx')
		
		break
	except IOError:
		print u'close the files'
print "saved buildings"
print "started trees"


p_final_tree={}

p_final_tree_roads={}
#Learning remaining relations for point elements
for pt_n in range(0,n_point):
	sheett=w_pt[pt_n].active#somethign p_n###################################################
	tccount=sheett.max_column-1
	t_gid=findattribute(sheett,u'gid')
	
	final_tree={}
	final_tree_roads={}
	
	tree_vectors=pt_vectors[pt_n]
	t_coordinates=pt_coordinates[pt_n]
	t_dist_coordinates=pt_dist_coordinates[pt_n]
	t_cartesian_coordinates=pt_cartesian_coordinates[pt_n]
	tree_dict=pt_dict[pt_n]

	for i in range(2,sheett.max_row+1):
		#print i
		print "Tree",i
		i_valid=[None,50]
		
		temp_t=t_coordinates[i-1][1]
		temp_t1=t_dist_coordinates[i-1]
		temp_t1=np.array(temp_t1)
		temp_c_t=t_cartesian_coordinates[i-1]
		#print t_coordinates[i-1][1][0]
		max_lat_allowed=float(t_coordinates[i-1][1][0])+float(.002)
		min_lat_allowed=float(t_coordinates[i-1][1][0])-float(.002)
		max_long_allowed=float(t_coordinates[i-1][1][1])+float(.002)
		min_long_allowed=float(t_coordinates[i-1][1][1])-float(.002)
		
		distances=[]
		indices=[]
		###########relations with linear elements#####################
		for b_n in range(0,n_block):
				
			tree_roads=b_seg_r_name[b_n]
			seg_dict=b_seg_dict[b_n]
			#road_dict=l_dict[b_n]
			road_vectors=l_vectors[b_n]
			sheets=w_b[b_n].active##
			s_gid_index=findattribute(sheets,u'gid')
		
			
				################### relation between building and road
			for l in range(0,len(tree_roads)):############################################################################################
	#if i==4:
			#	print "Length building roads= ",len(tree_roads),l, tree_roads
				jk=0
				road_segments=[]
				try:
				
					road_segments=seg_dict[tree_roads[l]] #[[i,temp2,seg_polygon,seg_polygon1,seg_c_polygon]]
				except KeyError:
					print "Segment of ", tree_roads[l], " not avalable for trees"
					continue
				
				min_distance_tree=pt_l_thresh
				seg_side=[]
				s_segment=[]
				flag111=0
				for n in range(0,len(road_segments)):
					
					segment=road_segments[n]
			
					for si in range(0,len(segment[4])-1):
						
						if segment[2][si][0] <= max_lat_allowed:
				
							if segment[2][si][0]>= min_lat_allowed:
								if segment[2][si][1]<= max_long_allowed:
									if segment[2][si][1]>= min_long_allowed:
					
										
										try:	
														
											if isInternal(temp_c_t,segment[4][si],segment[4][si+1]):
												t_seg_side=defineSideOfSegment(temp_c_t,segment[4][si],segment[4][si+1])
												result_tree=perDistancePointLine_2(temp_c_t,segment[4][si],segment[4][si+1])
												temp_min_distance_tree=haversine(cartesianToGeo(temp_c_t[0],temp_c_t[1]),cartesianToGeo(result_tree[1].x,result_tree[1].y))*1000
												if min_distance_tree>temp_min_distance_tree:
													seg_side=t_seg_side
													
													min_distance_tree=temp_min_distance_tree
													s_segment=segment
													flag111=1
										except Exception as e:
											print e
											continue
				
				temp_vector=tree_vectors[int(t_coordinates[i-1][0])]
					
				rand_1=temp_vector[0]
				rand_2=temp_vector[1]
				rand_3=temp_vector[2]
				n_pre=temp_vector[3]
				ave_pre=temp_vector[4]
				std_pre=temp_vector[5]
				n_new=n_pre+1
				ave_new=float((ave_pre*n_pre+min_distance_tree)/float(n_pre+1))
				sumx=ave_pre*n_pre
				sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
				std_new=(float((float(n_new)*(min_distance_tree**2+sumx2)-(min_distance_tree+sumx)**2)/float((n_new)**2)))**.5
				tree_vectors[int(t_coordinates[i-1][0])]=[rand_1,rand_2,rand_3,n_new,ave_new,std_new]
				
				
				#pdb.set_trace()
				temp_vector=road_vectors[int(tree_roads[l])]
					
				rand_1=temp_vector[0]
				rand_2=temp_vector[1]
				rand_3=temp_vector[2]
				n_pre=temp_vector[3]
				ave_pre=temp_vector[4]
				std_pre=temp_vector[5]
				n_new=n_pre+1
				ave_new=float((ave_pre*n_pre+min_distance_tree)/float(n_pre+1))
				sumx=ave_pre*n_pre
				sumx2=n_pre*(sumx*sumx+ave_pre*ave_pre)
				std_new=(float((float(n_new)*(min_distance_tree**2+sumx2)-(min_distance_tree+sumx)**2)/float((n_new)**2)))**.5
				road_vectors[int(tree_roads[l])]=[rand_1,rand_2,rand_3,n_new,ave_new,std_new]
				
				try:
					if flag111==1:
						final_tree_roads=p_final_tree_roads[pt_n]
						temp_final_roads=final_tree_roads[i]
						temp_final_roads.append([b_name[b_n]+str('_')+str(tree_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_tree,"Seg_Side_"+str(seg_side)])
						#tr_matrix[i][road_dict_2[road_dict[tree_roads[l]]]]=min_distance_tree
						final_tree_roads[i]=temp_final_roads
						p_final_tree_roads[pt_n]=final_tree_roads
					
				except KeyError:
					if flag111==1:
						temp_final_roads=[]
						temp_final_roads.append([b_name[b_n]+str('_')+str(tree_roads[l]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_tree,"Seg_Side_"+str(seg_side)])
						#tr_matrix[i][road_dict_2[road_dict[tree_roads[l]]]]=min_distance_tree
						final_tree_roads[i]=temp_final_roads
						p_final_tree_roads[pt_n]=final_tree_roads
						
		###########relations with point elements######################				
		for pt_n_j in range(pt_n,n_point):
			sheettj=w_pt[pt_n_j].active###
			tccount_j=sheett.max_column-1
			t_gid_j=findattribute(sheettj,u'gid')
									
			tree_vectors_j=pt_vectors[pt_n_j]
			t_coordinates_j=pt_coordinates[pt_n_j]
			t_dist_coordinates_j=pt_dist_coordinates[pt_n_j]
			t_cartesian_coordinates_j=pt_cartesian_coordinates[pt_n_j]
			tree_dict_j=pt_dict[pt_n]
			pt_s_index=2
			
			if pt_n==pt_n_j:
				pt_s_index=i+1
			final_tree_j={}
			if pt_s_index>sheettj.max_row:
				continue
			for j in range(pt_s_index,sheettj.max_row+1):
				
				l_lat=t_coordinates_j[j-1][1][0]
				l_long=t_coordinates_j[j-1][1][1]
				if l_lat <= max_lat_allowed:
					if l_lat>= min_lat_allowed:
						if l_long<= max_long_allowed:
							if l_long>= min_long_allowed:
								min_distance_tree=haversine((l_lat,l_long),(float(t_coordinates[i-1][1][0]),float(t_coordinates[i-1][1][1])))*1000
								if min_distance_tree<=pt_pt_thresh:
															
									try:
										final_tree=p_final_tree[pt_n]
										temp_final_tree=final_tree[i]
										temp_final_tree.append([pt_name[pt_n_j],str(t_coordinates_j[j-1][0]),min_distance_tree])
										final_tree[i]=temp_final_tree
										#bt_matrix[i][tree_dict[t_coordinates[potential_tree[jt][0]-1][0]]]=min_tree_distance
										p_final_tree[pt_n]=final_tree
									except KeyError:
										temp_final_tree=[]
										temp_final_tree.append([pt_name[pt_n_j],str(t_coordinates_j[j-1][0]),min_distance_tree])
										#bt_matrix[i][tree_dict[t_coordinates[potential_tree[jt][0]-1][0]]]=min_tree_distance
										final_tree[i]=temp_final_tree
										p_final_tree[pt_n]=final_tree
									try:
										final_tree_j=p_final_tree[pt_n_j]
										
										temp_final_tree=final_tree_j[j]
										temp_final_tree.append([pt_name[pt_n],str(t_coordinates[i-1][0]),min_distance_tree])
										final_tree_j[j]=temp_final_tree
										#bt_matrix[i][tree_dict[t_coordinates[potential_tree[jt][0]-1][0]]]=min_tree_distance
										p_final_tree[pt_n_j]=final_tree_j
										
									except KeyError:
										temp_final_tree=[]
										temp_final_tree.append([pt_name[pt_n],str(t_coordinates[i-1][0]),min_distance_tree])
										final_tree_j[j]=temp_final_tree
										p_final_tree[pt_n_j]=final_tree_j
										
							
			#print u'Done Trees',int(float(i)/float(sheett.max_row+1)*100),u'%'
		

print "done trees"
#Storing learned relations of polygon in workbook objects. The files are stored in working directory
for pt_n in range(0,n_point):
	sheett=w_pt[pt_n].active######
	final_tree=p_final_tree[pt_n]
	tccount=sheett.max_column
	for i in range(2,sheett.max_row+1):
		
		try:
			temp=final_tree[i]
			#print temp
			#temp=np.array(temp)
			#temp=temp[temp[:,2].argsort()]
			for pp in range (0,len(temp)):#([u'tree',u'Tree'+str(t_coordinates[potential_tree[j][0]-1][0]),min_tree_distance,"Side_"+str(a_building_side)])
				sheett.cell(row=1,column=tccount+3*(pp+1)-2).value=u'Object_Type'
				sheett.cell(row=1,column=tccount+3*(pp+1)-1).value=u'Object_Name'
				sheett.cell(row=1,column=tccount+3*(pp+1)).value=u'Object_Distance'	
		
				sheett.cell(row=i,column=tccount+3*(pp+1)-2).value=temp[pp][0]
				sheett.cell(row=i,column=tccount+3*(pp+1)-1).value=temp[pp][1]
				sheett.cell(row=i,column=tccount+3*(pp+1)).value=temp[pp][2]
				
		except KeyError:
			waste=1

for pt_n in range(0,n_point):
	sheett=w_pt[pt_n].active######
	final_tree_roads=p_final_tree_roads[pt_n]
	tccount=sheett.max_column
	for i in range(2,sheett.max_row+1):
		
		try:
			temp=final_tree_roads[i]
			#print temp
			#temp=np.array(temp)
			#temp=temp[temp[:,2].argsort()]
			for pp in range (0,len(temp)):#([u'tree',u'Tree'+str(t_coordinates[potential_tree[j][0]-1][0]),min_tree_distance,"Side_"+str(a_building_side)])
				sheett.cell(row=1,column=tccount+4*(pp+1)-3).value=u'Object_Type'
				sheett.cell(row=1,column=tccount+4*(pp+1)-2).value=u'Object_Name'
				sheett.cell(row=1,column=tccount+4*(pp+1)-1).value=u'Object_Distance'	
				sheett.cell(row=1,column=tccount+4*(pp+1)).value=u'Child_Side'
				sheett.cell(row=i,column=tccount+4*(pp+1)-3).value=temp[pp][0]
				sheett.cell(row=i,column=tccount+4*(pp+1)-2).value=temp[pp][1]
				sheett.cell(row=i,column=tccount+4*(pp+1)-1).value=temp[pp][2]
				sheett.cell(row=i,column=tccount+4*(pp+1)).value=temp[pp][3]
		except KeyError:
			waste=1
#code for storing workbook objects in disk. Each file is named with the corresponding element class (tree, light pole, traffic sensor) follwed by "-result"
while True:
	try:
		for pt_n in range(0,n_point):
			w_pt[pt_n].save(pt_name[pt_n]+'-result.xlsx')
		
		break
	except IOError:
		print u'close the files'
	
print "written trees"	
print "Starting Blocks"	
b_final_block={}
#Learning remaining relations of linear elements
for b_n in range(0,n_block):
	block_roads=b_seg_r_name[b_n]
	seg_dict=b_seg_dict[b_n]
	#road_dict=l_dict[b_n]
	sheets=w_b[b_n].active##
	s_gid_index=findattribute(sheets,u'gid')
	l_final_segments={}
	for m in range(0,len(block_roads)):############################################################################################
		print "Road " , m, " ", block_roads[m]
		jk=0
		road_segments=[]
		try:
			road_segments=seg_dict[block_roads[m]] #[[i,temp2,seg_polygon,seg_polygon1,seg_c_polygon]]
		except KeyError:
			print "Segment of ", block_roads[m], " not avalable for buildings"
			continue
		
		min_distance_segment=l_l_thresh
		seg_side=[]
		s_segment=[]
		
		for n in range(0,len(road_segments)):
			
			segment=road_segments[n]
			for si in range(0,len(segment[4])-1):
						
				max_lat_allowed=max(segment[2][si][0],segment[2][si+1][0])+float(.002)
				min_lat_allowed=min(segment[2][si][0],segment[2][si+1][0])-float(.002)
				max_long_allowed=max(segment[2][si][1],segment[2][si+1][1])+float(.002)
				min_long_allowed=min(segment[2][si][1],segment[2][si+1][1])-float(.002)
	###############################relations of linear elements with linear elements###################################
				for b_n_j in range(0,n_block):
					block_roads_j=b_seg_r_name[b_n_j]
					#road_dict_j=l_dict[b_n_j]
					seg_dict_j=b_seg_dict[b_n_j]
					sheetsj=w_b[b_n_j].active##
					s_gid_index_j=findattribute(sheetsj,u'gid')
					l_s_index=0
					if b_n==b_n_j:
						l_s_index=m+1
						
					if l_s_index>len(block_roads_j)-1:
						continue
					final_block={}
					for l in range(l_s_index,len(block_roads_j)):############################################################################################
			#if i==4:
					#	print "Length building roads= ",len(block_roads),l, block_roads
											
						jk=0
						road_segments_j=[]
						
						try:
						
							road_segments_j=seg_dict_j[block_roads_j[l]] #[[i,temp2,seg_polygon,seg_polygon1,seg_c_polygon]]
						except KeyError:
							print "Segment of ", block_roads_j[l], " not avalable for buildings"
							continue
						except IndexError:
							pdb.set_trace()
						
						min_distance_segment=l_l_thresh
						seg_side_j=[]
						s_segment_j=[]
						intersection_flag=0
						flag111=0
						for n_j in range(0,len(road_segments_j)):
							
							segment_j=road_segments_j[n_j]
					
							for si_j in range(0,len(segment_j[4])-1):
								
								if checkIfPointWithinBox([segment_j[2][si_j][0],segment_j[2][si_j+1][0]],[segment_j[2][si_j][1],segment_j[2][si_j+1][1]],max_lat_allowed,min_lat_allowed,max_long_allowed,min_long_allowed):
									if isInternal(segment_j[4][si_j],segment[4][si],segment[4][si+1]):
									
										if isInternal(segment_j[4][si_j+1],segment[4][si],segment[4][si+1]):
											side_of_i=defineSideOfSegment(segment_j[4][si_j+1],segment[4][si],segment[4][si+1])
											side_of_j=defineSideOfSegment(segment[4][si+1],segment_j[4][si_j],segment_j[4][si_j+1])
											temp_result_b=findShortestDistanceSegments(Point(segment[4][si][0],segment[4][si][1],0),Point(segment[4][si+1][0],segment[4][si+1][1],0),Point(segment_j[4][si_j][0],segment_j[4][si_j][1],0),Point(segment_j[4][si_j+1][0],segment_j[4][si_j+1][1],0))
											if temp_result_b[0][0]<0.001:
												intersection_flag=1
												min_distance_segment=0
												seg_side=0
												seg_side_j=0
												s_segment_j=segment_j
												break
											else:
												if min_distance_segment>temp_result_b[0][0]:
													min_distance_segment=temp_result_b[0][0]
													
													seg_side=side_of_i
													seg_side_j=side_of_j
													
													s_segment_j=segment_j
													flag111=1
								
							if intersection_flag==1:
								break
						if flag111==1:
						
							try:
							
								
								final_block=b_final_block[b_n]
								temp_final_block=final_block[segment[0]]
								temp_final_block.append([b_name[b_n_j]+str('_')+str(block_roads_j[l]),u'Seg'+str(sheetsj.cell(row=segment_j[0],column=s_gid_index_j).value),min_distance_segment,"Seg_Side_"+str(seg_side)])
								final_block[segment[0]]=temp_final_block
								b_final_block[b_n]=final_block
								
							except KeyError:
					
							
								temp_final_block=[]
								temp_final_block.append([b_name[b_n_j]+str('_')+str(block_roads_j[l]),u'Seg'+str(sheetsj.cell(row=segment_j[0],column=s_gid_index_j).value),min_distance_segment,"Seg_Side_"+str(seg_side)])
								final_block[segment[0]]=temp_final_block
								b_final_block[b_n]=final_block
							try:
							
								final_block=b_final_block[b_n_j]
								temp_final_block=final_block[segment_j[0]]
								temp_final_block.append([b_name[b_n]+str('_')+str(block_roads[m]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Seg_Side_"+str(seg_side_j)])
								final_block[segment[0]]=temp_final_block
								b_final_block[b_n_j]=final_block
								
							except KeyError:
								
								temp_final_block=[]
								temp_final_block.append([b_name[b_n]+str('_')+str(block_roads[m]),u'Seg'+str(sheets.cell(row=segment[0],column=s_gid_index).value),min_distance_segment,"Seg_Side_"+str(seg_side_j)])
								final_block[segment_j[0]]=temp_final_block
								b_final_block[b_n_j]=final_block

#Storing learned relations of polygon in workbook objects. The files are stored in working directory							
for b_n in range(0,n_block):
	sheets=w_b[b_n].active######
	final_block=b_final_block[pt_n]
	sccount=sheets.max_column
	for i in range(2,sheets.max_row+1):
		
		try:
			temp=final_block[i]
			#print temp
			#temp=np.array(temp)
			#temp=temp[temp[:,2].argsort()]
			for pp in range (0,len(temp)):#([u'tree',u'Tree'+str(t_coordinates[potential_tree[j][0]-1][0]),min_tree_distance,"Side_"+str(a_building_side)])
				sheets.cell(row=1,column=sccount+4*(pp+1)-3).value=u'Object_Type'
				sheets.cell(row=1,column=sccount+4*(pp+1)-2).value=u'Object_Name'
				sheets.cell(row=1,column=sccount+4*(pp+1)-1).value=u'Object_Distance'	
				sheets.cell(row=1,column=sccount+4*(pp+1)).value=u'Parent_Side'
				sheets.cell(row=i,column=sccount+4*(pp+1)-3).value=temp[pp][0]
				sheets.cell(row=i,column=sccount+4*(pp+1)-2).value=temp[pp][1]
				sheets.cell(row=i,column=sccount+4*(pp+1)-1).value=temp[pp][2]
				sheets.cell(row=i,column=sccount+4*(pp+1)).value=temp[pp][3]
		except KeyError:
			waste=1

#code for storing workbook objects in disk. Each file is named with the corresponding element class (tree, light pole, traffic sensor) follwed by "-result"
while True:
	try:
		for b_n in range(0,n_block):
			w_b[b_n].save(b_name[b_n]+'-result.xlsx')
		
		break
	except IOError:
		print u'close the files'				
								
print "Written Blocks"								

